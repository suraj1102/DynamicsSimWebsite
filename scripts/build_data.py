#!/usr/bin/env python3
"""Pre-render data pipeline for the Dynamics & Simulation course website.

Reads ``Dynamics&Sim-MasterSheet.xlsx`` and scans the material folders, then
emits JSON into ``_data/`` for the .qmd pages to render.

Nothing here hardcodes a year, an author, a lecture count or a problem count:
columns are detected by *pattern* (see CLAUDE.md §3), so adding a new video
year, notes-taker or solutions set is a spreadsheet + folder change only.

Run standalone for a verbose report:

    python3 scripts/build_data.py
"""

from __future__ import annotations

import datetime as _dt
import hashlib
import json
import re
import shutil
import sys
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "Dynamics&Sim-MasterSheet.xlsx"
DATA_DIR = ROOT / "_data"
FILES_DIR = ROOT / "files"
NOTES_ROOT = ROOT / "Notes"
HOMEWORK_ROOT = ROOT / "Homework"

# --- Column detection patterns (CLAUDE.md §3) -------------------------------
RE_LEC_NO = re.compile(r"^\s*lec\.?\s*no", re.I)
RE_LEC_NAME = re.compile(r"^\s*name\s+of\s+lecture", re.I)
RE_HW_NO = re.compile(r"^\s*homework\s*no", re.I)
RE_VIDEO_COL = re.compile(r"^\s*link\s*\((?P<label>[^)]*)\)\s*$", re.I)
RE_USAGE_COL = re.compile(r"^\s*done\s+last\s+time\s*\((?P<label>[^)]*)\)\s*$", re.I)

# --- Approval-status keywords (CLAUDE.md §4.3) ------------------------------
# Compared against filename + containing-folder name with all non-alphanumeric
# characters stripped and lowercased, so spacing/punctuation variants match.
# Edit this list to add new "not reviewed" markers.
UNAPPROVED_KEYWORDS = [
    "NotApproved",
    "NotYetApproved",
    "NotYetFinallyApproved",
    "NonFinalized",
    "NotAndyApproved",
    "NotAndyAproved",  # real misspelling in the 2017 set — keep
    "DamnGood",  # "NonApprovedButDamnGood" folder — implies not official
]
_UNAPPROVED_NORM = [re.sub(r"[^a-z0-9]", "", k.lower()) for k in UNAPPROVED_KEYWORDS]

# Files that are LaTeX build artifacts — never surfaced on the site (§8).
IGNORED_SUFFIXES = {".tex", ".sty", ".gz", ".aux", ".log", ".out", ".synctex"}
IGNORED_NAMES = {".DS_Store", "Thumbs.db"}

# The two PDFs copied into files/ under clean names, so the "#page=" deep links
# and the policy embed don't have to carry spaces and "&" in their URLs (§4.4).
PROBLEMS_PDF_SRC = "Homework/HomeworkAssignments/Homework Plaksha Dyn & Sim.pdf"
POLICY_PDF_SRC = "Homework/HomeworkAssignments/Homework-Policy.pdf"
PROBLEMS_PDF_HREF = "files/homework-problems.pdf"
POLICY_PDF_HREF = "files/homework-policy.pdf"
PDF_COPIES = {
    PROBLEMS_PDF_SRC: Path(PROBLEMS_PDF_HREF).name,
    POLICY_PDF_SRC: Path(POLICY_PDF_HREF).name,
}


LOG: list[str] = []


def warn(msg: str) -> None:
    LOG.append(msg)
    print(f"  [warn] {msg}", file=sys.stderr)


def is_ignored(path: Path) -> bool:
    return path.name in IGNORED_NAMES or path.suffix.lower() in IGNORED_SUFFIXES


def rel(path: Path) -> str:
    """Project-relative POSIX path, used as the href in the generated JSON."""
    return path.relative_to(ROOT).as_posix()


# ---------------------------------------------------------------------------
# Spreadsheet reading
# ---------------------------------------------------------------------------


def read_sheet(ws) -> tuple[list[str | None], list[tuple[int, list]]]:
    """Return (headers, rows) with blank rows and trailing blank columns dropped.

    Each row is ``(excel_row_number, values)``. The real row number is carried
    rather than recomputed from the list position, because blank rows are
    skipped — so position and worksheet row diverge as soon as the sheet has a
    gap in it. Warnings quote a cell coordinate the maintainer has to go and
    fix by hand, so it has to name the right cell.
    """
    headers = [c.value for c in ws[1]]
    while headers and headers[-1] in (None, ""):
        headers.pop()
    ncols = len(headers)

    rows = []
    for excel_row, row in enumerate(
        ws.iter_rows(min_row=2, max_col=ncols, values_only=True), start=2
    ):
        if all(v in (None, "") for v in row):
            continue
        rows.append((excel_row, list(row)))
    return headers, rows


def parse_note_numbers(value, where: str) -> tuple[list[int], str]:
    """Parse a notes-set cell into (note numbers, trailing commentary).

    Handles: blank, 7, 7.0, "16, 17", "03, 04 -> free text",
    "17, 18, 19 - free text". A datetime means the Excel-autocorrect bug from
    §1.1 survived — warn and skip rather than crash.
    """
    if value is None or value == "":
        return [], ""

    if isinstance(value, (_dt.datetime, _dt.date)):
        warn(
            f"{where}: cell holds a date ({value!r}) — this is the Excel "
            f"autocorrect bug (CLAUDE.md §1.1). Retype it as text. Skipping."
        )
        return [], ""

    if isinstance(value, (int, float)):
        return [int(value)], ""

    text = str(value).strip()
    # Leading run of digits/commas/whitespace is the number list; the rest is
    # free-text commentary (introduced by ->, -, :, — or nothing at all).
    m = re.match(r"^([0-9][0-9,\s]*)\s*(?:(?:->|[-–—:])\s*)?(\D.*)?$", text, re.S)
    if not m:
        warn(f"{where}: could not parse notes cell {text!r}. Skipping.")
        return [], ""

    numbers = [int(n) for n in re.findall(r"\d+", m.group(1))]
    commentary = (m.group(2) or "").strip()
    if not numbers:
        warn(f"{where}: no note numbers found in {text!r}. Skipping.")
    return numbers, commentary


def parse_urls(value) -> list[str]:
    """A video cell may hold one or several URLs, comma/whitespace separated.

    Split only where a separator is actually followed by the next URL, so a
    comma *inside* a URL (query strings legitimately contain them) does not
    truncate it.
    """
    if value is None:
        return []
    text = str(value).strip()
    parts = re.split(r"[,\s]+(?=https?://)", text)
    return [
        p.strip().rstrip(",")
        for p in parts
        if p.strip().startswith(("http://", "https://"))
    ]


def parse_bool(value) -> bool | None:
    if value is None or str(value).strip() == "":
        return None
    return str(value).strip().lower() in {"y", "yes", "true", "1"}


# ---------------------------------------------------------------------------
# Notes folders
# ---------------------------------------------------------------------------


def leading_note_numbers(filename: str) -> list[int]:
    """Note number(s) a notes filename declares in its leading position.

    The two existing sets name files differently, and the resolution rule has
    to cover both without knowing the author (CLAUDE.md §3):

        "1. 08-24 Intro-pos-vel.pdf"          -> [1]
        "12 09-21 Intro to Constraints.pdf"   -> [12]
        "Lecture 01 Introduction.pdf"         -> [1]
        "Lecture 09 10 Equilibrium.pdf"       -> [9, 10]   (one file, two notes)
        "39b. 11-30 Ref Frames Picture.jpg"   -> []        (not a plain number)

    Leading purely-alphabetic words ("Lecture") are skipped; the first token
    containing a digit must be a plain number or matching stops, which keeps
    embedded dates like "08-24" from being read as note numbers.
    """
    stem = Path(filename).stem
    numbers: list[int] = []
    for token in stem.split():
        clean = token.rstrip(".")
        if clean.isdigit():
            numbers.append(int(clean))
            continue
        if numbers:
            break  # end of the leading number run
        if any(ch.isdigit() for ch in token):
            break  # e.g. "39b." or "08-24" — not a plain note number
        continue  # purely-alphabetic prefix word, e.g. "Lecture"
    return numbers


def index_notes_folder(folder: Path) -> dict[int, list[Path]]:
    index: dict[int, list[Path]] = {}
    for path in sorted(folder.iterdir()):
        if not path.is_file() or is_ignored(path):
            continue
        for n in leading_note_numbers(path.name):
            index.setdefault(n, []).append(path)
    return index


# ---------------------------------------------------------------------------
# Solutions folders
# ---------------------------------------------------------------------------


def is_unapproved(*names: str) -> bool:
    for name in names:
        norm = re.sub(r"[^a-z0-9]", "", name.lower())
        if any(k in norm for k in _UNAPPROVED_NORM):
            return True
    return False


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


# The tidy "Q##_COURSE_TERM[_more].pdf" convention, preferred when duplicates
# exist (§4.3) over looser names like "Q45 - NotAndyApproved.pdf".
CLEAN_SOLUTION_NAME = re.compile(r"^Q\d+(?:_[A-Za-z0-9]+){2,}\.pdf$", re.I)


def dedupe_preference(path: Path) -> tuple:
    """Sort key picking the survivor among byte-identical duplicates (§4.3):
    the cleaner ``Q##_COURSE_TERM.pdf`` name wins, then the shallower/shorter path."""
    return (
        0 if CLEAN_SOLUTION_NAME.match(path.name) else 1,
        len(path.relative_to(ROOT).parts),
        len(path.name),
        path.name,
    )


@dataclass
class SolutionSet:
    label: str
    folder: Path
    # every candidate file, top level + one subfolder level deep
    files: list[Path] = field(default_factory=list)
    consumed: set[Path] = field(default_factory=set)

    def scan(self) -> None:
        for path in sorted(self.folder.iterdir()):
            if path.is_file() and not is_ignored(path):
                self.files.append(path)
            elif path.is_dir():
                for child in sorted(path.iterdir()):
                    if child.is_file() and not is_ignored(child):
                        self.files.append(child)

    def match(self, code: str, where: str) -> dict | None:
        """Resolve a solutions-cell code (e.g. "Q01") to a PDF + supporting files."""
        # The lookahead stops code "Q1" from matching "Q12_....pdf", while still
        # allowing a bare "Q01.pdf" (which a trailing \D would reject).
        code_re = re.compile(rf"^{re.escape(code)}(?!\d).*\.pdf$", re.I)
        # A subfolder named for the code carries the supporting .m/.png files.
        subfolders = {
            p.parent
            for p in self.files
            if p.parent != self.folder
            and re.match(rf"^{re.escape(code)}(?!\d)", p.parent.name, re.I)
        }

        candidates = [
            p
            for p in self.files
            if code_re.match(p.name) and (p.parent == self.folder or p.parent in subfolders)
        ]

        supporting = sorted(
            p for sub in subfolders for p in self.files if p.parent == sub
        )

        if not candidates:
            if supporting:
                warn(
                    f"{where}: no PDF matching {code!r} in {self.label!r}; "
                    f"falling back to the {len(supporting)} file(s) in its subfolder."
                )
                for p in supporting:
                    self.consumed.add(p)
                folder_names = [s.name for s in subfolders]
                return {
                    "set": self.label,
                    "code": code,
                    "file": None,
                    "approved": not is_unapproved(*folder_names),
                    "supporting_files": [
                        {"name": p.name, "file": rel(p)} for p in supporting
                    ],
                }
            warn(f"{where}: no file matching {code!r} in {self.label!r}.")
            return None

        chosen = min(dedupe(candidates), key=dedupe_preference)
        for p in candidates:
            self.consumed.add(p)

        # Everything else in the code's subfolder rides along as a supporting
        # file. Candidates are excluded: the chosen one is the solution itself,
        # and the others are its byte-identical duplicates.
        supporting_out = []
        for p in supporting:
            if p in candidates:
                continue
            self.consumed.add(p)
            supporting_out.append({"name": p.name, "file": rel(p)})

        return {
            "set": self.label,
            "code": code,
            "file": rel(chosen),
            "approved": not is_unapproved(chosen.name, chosen.parent.name),
            "supporting_files": supporting_out,
        }

    def leftovers(self) -> list[dict]:
        """Files in the folder no sheet row referenced — the §4.3 "other materials"."""
        rest = [p for p in self.files if p not in self.consumed]
        out = []
        for p in sorted(dedupe(rest), key=lambda q: q.name.lower()):
            out.append(
                {
                    "name": p.stem,
                    "file": rel(p),
                    "subfolder": p.parent.name if p.parent != self.folder else None,
                    "approved": not is_unapproved(p.name, p.parent.name),
                }
            )
        return out


def dedupe(paths: list[Path]) -> list[Path]:
    """Drop byte-identical duplicates, keeping the preferred name (§4.3)."""
    by_hash: dict[str, list[Path]] = {}
    for p in paths:
        by_hash.setdefault(sha256(p), []).append(p)
    kept = []
    for group in by_hash.values():
        group.sort(key=dedupe_preference)
        kept.append(group[0])
        for dropped in group[1:]:
            LOG.append(f"deduped {rel(dropped)} (identical to {rel(group[0])})")
    return kept


# ---------------------------------------------------------------------------
# Homework problem -> PDF page mapping (§4.4)
# ---------------------------------------------------------------------------


def problem_pages(pdf: Path, hw_numbers: list[int]) -> dict[int, int]:
    """Map each homework number to the PDF page its "N. Title" line starts on.

    Search is monotonic — problem N is looked for at or after the page where
    N-1 was found — so stray numbers (equation labels, part numbers) elsewhere
    in the document cannot capture a match.
    """
    reader = PdfReader(str(pdf))
    pages = [(p.extract_text() or "") for p in reader.pages]

    found: dict[int, int] = {}
    start = 0
    for n in sorted(hw_numbers):
        pattern = re.compile(rf"^[ \t]*{n}\.[ \t]+\S", re.M)
        for i in range(start, len(pages)):
            if pattern.search(pages[i]):
                found[n] = i + 1
                start = i
                break
    return found


def copy_clean_pdfs() -> None:
    FILES_DIR.mkdir(exist_ok=True)
    for src_rel, dest_name in PDF_COPIES.items():
        src = ROOT / src_rel
        if not src.exists():
            warn(f"expected PDF missing: {src_rel}")
            continue
        shutil.copy2(src, FILES_DIR / dest_name)


# ---------------------------------------------------------------------------
# Build
# ---------------------------------------------------------------------------


def build_lectures(ws) -> tuple[list[dict], list[dict], list[str]]:
    headers, rows = read_sheet(ws)

    no_col = name_col = None
    video_cols: list[tuple[int, str]] = []
    notes_cols: list[tuple[int, str]] = []

    for i, h in enumerate(headers):
        if h is None or str(h).strip() == "":
            continue
        h = str(h).strip()
        if no_col is None and RE_LEC_NO.match(h):
            no_col = i
        elif name_col is None and RE_LEC_NAME.match(h):
            name_col = i
        elif (m := RE_VIDEO_COL.match(h)):
            video_cols.append((i, m.group("label").strip()))
        else:
            notes_cols.append((i, h))

    if no_col is None or name_col is None:
        raise SystemExit("Lectures sheet: missing 'Lec No.' / 'Name of Lecture' column.")

    # Each notes column header must name a folder under Notes/ exactly (§3).
    notes_index: dict[str, dict[int, list[Path]]] = {}
    for _, label in notes_cols:
        folder = NOTES_ROOT / label
        if not folder.is_dir():
            warn(
                f"Lectures sheet has notes column {label!r} but no folder "
                f"'Notes/{label}'. Create it (name must match exactly) — "
                f"skipping this set."
            )
            continue
        notes_index[label] = index_notes_folder(folder)

    lectures = []
    for excel_row, row in rows:
        if row[no_col] is None:
            continue
        lec_no = int(row[no_col])
        entry = {
            "lec_no": lec_no,
            "name": str(row[name_col] or "").strip(),
            "videos": [],
            "notes": [],
        }

        for i, label in video_cols:
            for url in parse_urls(row[i]):
                entry["videos"].append({"label": label, "url": url})

        for i, label in notes_cols:
            if label not in notes_index:
                continue
            where = (
                f"Lectures!{ws.cell(row=excel_row, column=i + 1).coordinate} "
                f"(lecture {lec_no}, {label})"
            )
            numbers, commentary = parse_note_numbers(row[i], where)
            for n in numbers:
                matches = notes_index[label].get(n)
                if not matches:
                    warn(f"{where}: no file for note {n} in 'Notes/{label}'.")
                    continue
                for path in matches:
                    # One file can cover several note numbers ("Lecture 09 10
                    # ....pdf"); show it once, listing every number it covers.
                    href = rel(path)
                    existing = next(
                        (
                            it
                            for it in entry["notes"]
                            if it["set"] == label and it["file"] == href
                        ),
                        None,
                    )
                    if existing:
                        if n not in existing["note_nos"]:
                            existing["note_nos"].append(n)
                        continue
                    item = {
                        "set": label,
                        "note_no": n,
                        "note_nos": [n],
                        "file": href,
                        "title": path.stem,
                    }
                    if commentary:
                        item["note"] = commentary
                    entry["notes"].append(item)

        for item in entry["notes"]:
            item["note_nos"].sort()
            item["note_no"] = item["note_nos"][0]
        entry["notes"].sort(key=lambda it: (it["set"], it["note_no"]))

        lectures.append(entry)

    video_labels = [label for _, label in video_cols]
    notes_labels = [label for _, label in notes_cols if label in notes_index]
    return lectures, video_labels, notes_labels


def build_homeworks(ws) -> tuple[list[dict], dict]:
    headers, rows = read_sheet(ws)

    no_col = None
    usage_cols: list[tuple[int, str]] = []
    solution_cols: list[tuple[int, str]] = []

    for i, h in enumerate(headers):
        if h is None or str(h).strip() == "":
            continue
        h = str(h).strip()
        if no_col is None and RE_HW_NO.match(h):
            no_col = i
        elif (m := RE_USAGE_COL.match(h)):
            usage_cols.append((i, m.group("label").strip()))
        else:
            solution_cols.append((i, h))

    if no_col is None:
        raise SystemExit("Homeworks sheet: missing 'Homework No. (Plaksha Doc)' column.")

    sets: dict[str, SolutionSet] = {}
    for _, label in solution_cols:
        folder = HOMEWORK_ROOT / label
        if not folder.is_dir():
            warn(
                f"Homeworks sheet has solutions column {label!r} but no folder "
                f"'Homework/{label}'. Create it (name must match exactly) — "
                f"skipping this set."
            )
            continue
        s = SolutionSet(label=label, folder=folder)
        s.scan()
        sets[label] = s

    hw_numbers = [int(row[no_col]) for _, row in rows if row[no_col] is not None]
    pages = problem_pages(ROOT / PROBLEMS_PDF_SRC, hw_numbers)
    unmatched = [n for n in hw_numbers if n not in pages]
    if unmatched:
        warn(f"homework problems with no PDF page match (linking whole doc): {unmatched}")

    homeworks = []
    for excel_row, row in rows:
        if row[no_col] is None:
            continue
        hw_no = int(row[no_col])
        entry = {
            "hw_no": hw_no,
            "usage": [],
            "problem_link": {
                "pdf": PROBLEMS_PDF_HREF,
                "page": pages.get(hw_no, 1),
                "matched": hw_no in pages,
            },
            "solutions": [],
        }

        for i, label in usage_cols:
            done = parse_bool(row[i])
            if done is not None:
                entry["usage"].append({"label": label, "done": done})

        for i, label in solution_cols:
            if label not in sets or row[i] is None or str(row[i]).strip() == "":
                continue
            code = str(row[i]).strip()
            where = (
                f"Homeworks!{ws.cell(row=excel_row, column=i + 1).coordinate} "
                f"(problem {hw_no}, {label})"
            )
            match = sets[label].match(code, where)
            if match:
                entry["solutions"].append(match)

        homeworks.append(entry)

    other = {label: s.leftovers() for label, s in sets.items()}
    meta = {
        "usage_labels": [label for _, label in usage_cols],
        "solution_sets": list(sets.keys()),
        "other_materials": other,
    }
    return homeworks, meta


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"master sheet not found: {XLSX}")

    DATA_DIR.mkdir(exist_ok=True)
    copy_clean_pdfs()

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    lectures, video_sets, notes_sets = build_lectures(wb["Lectures"])
    homeworks, hw_meta = build_homeworks(wb["Homeworks"])

    meta = {
        "video_sets": video_sets,
        "notes_sets": notes_sets,
        "usage_labels": hw_meta["usage_labels"],
        "solution_sets": hw_meta["solution_sets"],
        "counts": {"lectures": len(lectures), "homeworks": len(homeworks)},
        "policy_pdf": POLICY_PDF_HREF,
        "problems_pdf": PROBLEMS_PDF_HREF,
        "warnings": LOG,
    }

    for name, payload in (
        ("lectures.json", lectures),
        ("homeworks.json", homeworks),
        ("other_materials.json", hw_meta["other_materials"]),
        ("meta.json", meta),
    ):
        (DATA_DIR / name).write_text(
            json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8"
        )

    print(
        f"build_data: {len(lectures)} lectures, {len(homeworks)} homework problems, "
        f"video sets {meta['video_sets']}, notes sets {notes_sets}, "
        f"solutions sets {hw_meta['solution_sets']}, {len(LOG)} note(s)."
    )


if __name__ == "__main__":
    main()
